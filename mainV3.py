import PySimpleGUI as sg
import openpyxl
import logging
import getpass
import os
from datetime import datetime

# ----------------------------------------------------------------
# CONFIGURAÇÕES INICIAIS
# ----------------------------------------------------------------

# Caminhos das bases de dados (mesmo que você já tinha)
caminho_base_estoque      = r'\\nas1\infra\001 - Base Controle Estoque\estoque.xlsx'
caminho_registro_retiradas = r'\\nas1\infra\001 - Base Controle Estoque\registro_retiradas.xlsx'
caminho_registro_entradas  = r'\\nas1\infra\001 - Base Controle Estoque\registro_entradas.xlsx'

diretorio_base_estoque = os.path.dirname(caminho_base_estoque)
if not os.path.isdir(diretorio_base_estoque):
    os.makedirs(diretorio_base_estoque, exist_ok=True)

# Configuração de logging
logging.basicConfig(
    filename=os.path.join(diretorio_base_estoque, 'estoque_log.txt'),
    level=logging.INFO,
    format='%(asctime)s - %(message)s'
)

def log_acao(usuario, acao, descricao, quantidade=None, id_chamado=None,
             responsavel=None, numeros_series=None, numero_pedido=None, area_cc=None):
    """Grava ação no arquivo de log."""
    if quantidade is not None:
        logging.info(
            f"{usuario} {acao} '{descricao}' com quantidade {quantidade}. "
            f"ID Chamado: {id_chamado}, Responsável: {responsavel}, "
            f"Números de Série: {numeros_series}, Número Pedido: {numero_pedido}, Área/CC: {area_cc}"
        )
    else:
        logging.info(f"{usuario} {acao} '{descricao}'")

# Usuários autorizados (mesmos do seu código original)
usuarios_autorizados = {
    'lucas.gomes': '',
    'lucas.reis.adm': '',
    'juan.soares': '',
    'juan.soares.adm': '',
    'luan.freitas': '',
    'luan.freitas.adm': '',
    'rafael.andrade': '',
    'rafael.andrade.adm': '',
    'victor.gomes': '',
    'victor.gomes.adm': '',
    'lucelio.pereira': '',
    'lucelio.pereira.adm': '',
    'antonio.caroba': '',
    'antonio.caroba.adm': ''
}

usuario_logado = getpass.getuser()
if usuario_logado not in usuarios_autorizados:
    sg.popup_error("Usuário não autorizado!")
    exit()

# ----------------------------------------------------------------
# FUNÇÕES DE NEGÓCIO (idem ao seu script original, sem alteração)
# ----------------------------------------------------------------

def atualizar_estoque(descricao, quantidade):
    wb = openpyxl.load_workbook(caminho_base_estoque)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == descricao:
            nova_quantidade = row[1] - quantidade
            ws.cell(row=i + 1, column=2, value=nova_quantidade)
            break

    wb.save(caminho_base_estoque)

def verificar_id_chamado(id_chamado):
    if not os.path.exists(caminho_registro_retiradas):
        return False

    wb = openpyxl.load_workbook(caminho_registro_retiradas)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if row[4] == id_chamado:
            return True
    return False

def baixar_estoque(usuario, descricao, quantidade, id_chamado=None, responsavel=None, numeros_series=None):
    if verificar_id_chamado(id_chamado):
        sg.popup_error("ID de chamado já registrado. Use um ID único.")
        return

    numeros_series_list = numeros_series.split(',')
    if len(numeros_series_list) != quantidade:
        sg.popup_error("A quantidade de números de série não corresponde à quantidade de itens.")
        return

    wb = openpyxl.load_workbook(caminho_base_estoque)
    ws = wb.active

    item_encontrado = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == descricao:
            item_encontrado = True
            if row[1] >= quantidade:
                nova_quantidade = row[1] - quantidade
                ws.cell(row=i + 1, column=2, value=nova_quantidade)
                wb.save(caminho_base_estoque)
                registrar_retirada(usuario, descricao, quantidade, id_chamado, responsavel, numeros_series)
                sg.popup_ok(f"{quantidade} unidades de '{descricao}' foram retiradas do estoque.")
                log_acao(usuario, "RETIROU", descricao, quantidade, id_chamado, responsavel, numeros_series)
            else:
                sg.popup_error(f"Quantidade insuficiente de '{descricao}' em estoque.")
            break

    if not item_encontrado:
        sg.popup_error(f"Item '{descricao}' não encontrado no estoque.")

def registrar_retirada(usuario, descricao, quantidade, id_chamado, responsavel, numeros_series):
    if not os.path.exists(caminho_registro_retiradas):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Data", "Usuário", "Descrição", "Quantidade", "ID Chamado", "Responsável", "Números de Série"])
    else:
        wb = openpyxl.load_workbook(caminho_registro_retiradas)
        ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        usuario, descricao, quantidade, id_chamado, responsavel, numeros_series
    ])
    wb.save(caminho_registro_retiradas)

def cadastrar_equipamento(usuario, descricao, quantidade, numero_pedido, area_cc):
    wb = openpyxl.load_workbook(caminho_base_estoque)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if row[0] == descricao:
            sg.popup_error(f"O equipamento '{descricao}' já está cadastrado!")
            return

    ws.append([descricao, quantidade])
    wb.save(caminho_base_estoque)

    registrar_entrada(usuario, descricao, quantidade, numero_pedido, area_cc)
    sg.popup_ok(f"Equipamento '{descricao}' cadastrado com sucesso!")
    log_acao(usuario, "CADASTROU", descricao, quantidade)

def adicionar_estoque(usuario, descricao, quantidade, numero_pedido, area_cc):
    wb = openpyxl.load_workbook(caminho_base_estoque)
    ws = wb.active

    item_encontrado = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == descricao:
            nova_quantidade = row[1] + quantidade
            ws.cell(row=i + 1, column=2, value=nova_quantidade)
            item_encontrado = True
            break

    if not item_encontrado:
        sg.popup_error(f"Item '{descricao}' não encontrado no estoque.")
        return

    wb.save(caminho_base_estoque)
    registrar_entrada(usuario, descricao, quantidade, numero_pedido, area_cc)
    sg.popup_ok(f"Quantidade de '{descricao}' aumentada em {quantidade} unidades!")
    log_acao(usuario, "ADICIONOU", descricao, quantidade, numero_pedido=numero_pedido, area_cc=area_cc)

def registrar_entrada(usuario, descricao, quantidade, numero_pedido, area_cc):
    if not os.path.exists(caminho_registro_entradas):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Data", "Usuário", "Descrição", "Quantidade", "Número Pedido", "Área/CC"])
    else:
        wb = openpyxl.load_workbook(caminho_registro_entradas)
        ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        usuario, descricao, quantidade, numero_pedido, area_cc
    ])
    wb.save(caminho_registro_entradas)

def consultar_estoque(descricao):
    wb = openpyxl.load_workbook(caminho_base_estoque)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if row[0] == descricao:
            sg.popup_ok(f"Estoque de '{descricao}': {row[1]} unidades")
            return

def consultar_registros(tipo):
    """Mostra uma janela com a tabela de registros de entradas ou retiradas."""
    # Define caminho e cabeçalhos dependendo do tipo
    if tipo == 'entradas':
        caminho = caminho_registro_entradas
        headers = ["Data", "Usuário", "Descrição", "Quantidade", "Número Pedido", "Área/CC"]
    else:
        caminho = caminho_registro_retiradas
        headers = ["Data", "Usuário", "Descrição", "Quantidade", "ID Chamado", "Responsável", "Números de Série"]

    # Verifica se o arquivo existe
    if not os.path.exists(caminho):
        sg.popup_error(f"Não há registros de {tipo}.")
        return

    # Abre a planilha e pega a aba ativa
    wb = openpyxl.load_workbook(caminho)
    ws = wb.active

    # CORREÇÃO: usamos values_only=True e cada `row` já é uma tupla de valores (sem objetos Cell)
    registros = [ tuple(row) for row in ws.iter_rows(values_only=True) ]

    # Montagem do layout da janela de exibição
    layout = [
        [
            sg.Table(
                values=registros,
                headings=headers,
                auto_size_columns=True,
                justification='center',
                num_rows=min(len(registros), 20),
                font=('Helvetica', 11),
                alternating_row_color='#2A2A2A',
                background_color='#1E1E1E',
                header_background_color='#555555',
                header_text_color='white',
                text_color='white',
                key='-TABLE_REG-',
                enable_events=False
            )
        ],
        [sg.Push(), sg.Button("Fechar", size=(10,1)), sg.Push()]
    ]

    janela = sg.Window(f"Registros de {tipo.capitalize()}", layout, modal=True, finalize=True)
    while True:
        ev, _ = janela.read()
        if ev in (sg.WIN_CLOSED, 'Fechar'):
            break
    janela.close()

def deletar_equipamento(usuario, descricao):
    wb = openpyxl.load_workbook(caminho_base_estoque)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == descricao:
            ws.delete_rows(i + 1)
            wb.save(caminho_base_estoque)
            sg.popup_ok(f"Equipamento '{descricao}' deletado com sucesso!")
            log_acao(usuario, "DELETOU", descricao)
            return

    sg.popup_error(f"Equipamento '{descricao}' não encontrado!")

def carregar_lista_descricoes():
    """Retorna lista de descrições existentes no estoque."""
    if not os.path.exists(caminho_base_estoque):
        return []
    wb = openpyxl.load_workbook(caminho_base_estoque)
    ws = wb.active
    return [row[0] for row in ws.iter_rows(values_only=True)]

def coletar_dados_estoque():
    """Retorna lista de [descricao, quantidade] para tabela."""
    if not os.path.exists(caminho_base_estoque):
        return [], 0
    wb = openpyxl.load_workbook(caminho_base_estoque)
    ws = wb.active
    dados = []
    total_itens = 0
    for row in ws.iter_rows(values_only=True):
        dados.append([row[0], row[1]])
        total_itens += row[1]
    return dados, total_itens

# ----------------------------------------------------------------
# LAYOUT PRINCIPAL (com abas)
# ----------------------------------------------------------------

sg.theme('DarkTeal9')
sg.set_options(font=('Helvetica', 12), element_padding=(6, 4))

# Aba “Visão Geral” (tabela e total)
aba_visao = sg.Tab(
    "📊 Visão Geral",
    [
        [
            sg.Text(f"Bem-vindo, {usuario_logado}", font=("Helvetica", 20, "bold"), justification="center", expand_x=True)
        ],
        [sg.HSep()],
        [
            sg.Table(
                values=[],
                headings=["Descrição", "Quantidade"],
                key='-TABELA-',
                auto_size_columns=False,
                col_widths=[40, 12],
                justification='center',
                num_rows=15,
                font=('Consolas', 11),
                alternating_row_color='#2A2A2A',
                background_color='#1E1E1E',
                header_background_color='#555555',
                header_text_color='white',
                text_color='white',
                expand_x=True,
                expand_y=True
            )
        ],
        [
            sg.Text("Total de itens em estoque: 0", key='-TOTAL-', font=("Helvetica", 16, "bold"), justification="center", expand_x=True)
        ],
        [sg.Button("🔄 Atualizar", key='-ATUALIZA-VISUAL-', size=(12,1))]
    ]
)

# Aba “Entrada” (Adicionar quantidade)
descricoes_combo_entrada = carregar_lista_descricoes()
aba_entrada = sg.Tab(
    "➕ Entrada",
    [
        [sg.Text("Escolha o equipamento para adicionar ao estoque:", font=("Helvetica", 14))],
        [sg.Combo(descricoes_combo_entrada, key='ent_descr', size=(40,1)), sg.Stretch()],
        [sg.Text("Quantidade:", size=(12,1)), sg.Input(key='ent_qt', size=(10,1)), sg.Stretch()],
        [sg.Text("Número do Pedido:", size=(12,1)), sg.Input(key='ent_pedido', size=(25,1)), sg.Stretch()],
        [sg.Text("Área/CC:", size=(12,1)), sg.Input(key='ent_area', size=(25,1)), sg.Stretch()],
        [
            sg.Button("💾 Adicionar", key='-BOTAO-ENTRADA-', size=(12,1)),
            sg.Button("✖ Cancelar", key='-CANCELA-ENTRADA-', size=(12,1))
        ]
    ]
)

# Aba “Saída” (Baixar estoque)
descricoes_combo_saida = carregar_lista_descricoes()
aba_saida = sg.Tab(
    "➖ Saída",
    [
        [sg.Text("Escolha o equipamento para retirar do estoque:", font=("Helvetica", 14))],
        [sg.Combo(descricoes_combo_saida, key='sai_descr', size=(40,1)), sg.Stretch()],
        [sg.Text("Quantidade:", size=(12,1)), sg.Input(key='sai_qt', size=(10,1)), sg.Stretch()],
        [sg.Text("ID do Chamado:", size=(12,1)), sg.Input(key='sai_id', size=(25,1)), sg.Stretch()],
        [sg.Text("Responsável:", size=(12,1)), sg.Input(key='sai_resp', size=(25,1)), sg.Stretch()],
        [sg.Text("Números de Série (vírgula):", size=(18,1)), sg.Input(key='sai_numser', size=(30,1)), sg.Stretch()],
        [
            sg.Button("💾 Baixar", key='-BOTAO-SAIDA-', size=(12,1)),
            sg.Button("✖ Cancelar", key='-CANCELA-SAIDA-', size=(12,1))
        ]
    ]
)

# Aba “Cadastro” (Cadastrar novo equipamento)
aba_cadastro = sg.Tab(
    "🆕 Cadastro",
    [
        [sg.Text("Descrição do Equipamento:", size=(18,1)), sg.Input(key='cad_descr', size=(30,1)), sg.Stretch()],
        [sg.Text("Quantidade inicial:", size=(18,1)), sg.Input(key='cad_qt', size=(10,1)), sg.Stretch()],
        [sg.Text("Número do Pedido:", size=(18,1)), sg.Input(key='cad_pedido', size=(25,1)), sg.Stretch()],
        [sg.Text("Área/CC:", size=(18,1)), sg.Input(key='cad_area', size=(25,1)), sg.Stretch()],
        [
            sg.Button("💾 Cadastrar", key='-BOTAO-CADASTRO-', size=(12,1)),
            sg.Button("✖ Cancelar", key='-CANCELA-CADASTRO-', size=(12,1))
        ]
    ]
)

# Aba “Exclusão” (Deletar equipamento)
descricoes_combo_del = carregar_lista_descricoes()
aba_exclusao = sg.Tab(
    "🗑 Exclusão",
    [
        [sg.Text("Selecione o equipamento para deletar:", font=("Helvetica", 14))],
        [sg.Listbox(descricoes_combo_del, size=(40, 10), key='del_descr', enable_events=False)],
        [
            sg.Button("🗑 Deletar", key='-BOTAO-DEL-', size=(12,1)),
            sg.Button("✖ Cancelar", key='-CANCELA-DEL-', size=(12,1))
        ]
    ]
)

# Aba “Registros” (Entradas / Saídas)
aba_registros = sg.Tab(
    "📋 Registros",
    [
        [sg.Text("Consultar registros de estoque:", font=("Helvetica", 14))],
        [
            sg.Button("📥 Entradas", key='-BOTAO-REG-ENT-', size=(12,1)),
            sg.Button("📤 Saídas", key='-BOTAO-REG-SAI-', size=(12,1))
        ]
    ]
)

layout = [
    [sg.TabGroup(
        [[aba_visao, aba_entrada, aba_saida, aba_cadastro, aba_exclusao, aba_registros]],
        tab_location='centertop',
        selected_title_color='white',
        background_color='#1E1E1E',
        tab_background_color='#2A2A2A',
        border_width=2,
        title_color='lightgray',
        key='-TABGROUP-',
        expand_x=True,
        expand_y=True
    )]
]

window = sg.Window(
    "Controle de Estoque • Versão Moderna",
    layout,
    size=(800, 600),
    resizable=True,
    finalize=True,
    element_justification='center'
)

# Preenche tabela e total logo ao iniciar
def refresh_tabela_total():
    dados, total = coletar_dados_estoque()
    window['-TABELA-'].update(values=dados)
    window['-TOTAL-'].update(f"Total de itens em estoque: {total}")

refresh_tabela_total()

# ----------------------------------------------------------------
# LOOP PRINCIPAL
# ----------------------------------------------------------------

while True:
    event, values = window.read()
    if event in (sg.WIN_CLOSED,):
        break

    # ----------------------------------------------------------------
    # ABA “Visão Geral”
    # ----------------------------------------------------------------
    if event == '-ATUALIZA-VISUAL-':
        refresh_tabela_total()

    # ----------------------------------------------------------------
    # ABA “Entrada”
    # ----------------------------------------------------------------
    elif event == '-BOTAO-ENTRADA-':
        desc = values.get('ent_descr')
        try:
            qt = int(values.get('ent_qt'))
        except (ValueError, TypeError):
            sg.popup_error("Informe uma quantidade válida (número inteiro).")
            continue
        num_pedido = values.get('ent_pedido')
        area_cc    = values.get('ent_area')
        if not desc:
            sg.popup_error("Selecione um equipamento.")
        else:
            adicionar_estoque(usuario_logado, desc, qt, num_pedido, area_cc)
            # Atualiza combo das outras abas e tabela
            descricoes_combo_entrada = carregar_lista_descricoes()
            descricoes_combo_saida   = carregar_lista_descricoes()
            window['ent_descr'].update(values=descricoes_combo_entrada)
            window['sai_descr'].update(values=descricoes_combo_saida)
            window['del_descr'].update(values=descricoes_combo_entrada)
            refresh_tabela_total()

    elif event == '-CANCELA-ENTRADA-':
        window['ent_descr'].update(value='')
        window['ent_qt'].update(value='')
        window['ent_pedido'].update(value='')
        window['ent_area'].update(value='')

    # ----------------------------------------------------------------
    # ABA “Saída”
    # ----------------------------------------------------------------
    elif event == '-BOTAO-SAIDA-':
        desc = values.get('sai_descr')
        try:
            qt = int(values.get('sai_qt'))
        except (ValueError, TypeError):
            sg.popup_error("Informe uma quantidade válida (número inteiro).")
            continue
        id_cham = values.get('sai_id')
        resp    = values.get('sai_resp')
        num_ser = values.get('sai_numser')
        if not desc:
            sg.popup_error("Selecione um equipamento.")
        else:
            baixar_estoque(usuario_logado, desc, qt, id_cham, resp, num_ser)
            # Atualiza combo e tabela
            descricoes_combo_entrada = carregar_lista_descricoes()
            descricoes_combo_saida   = carregar_lista_descricoes()
            window['ent_descr'].update(values=descricoes_combo_entrada)
            window['sai_descr'].update(values=descricoes_combo_saida)
            window['del_descr'].update(values=descricoes_combo_entrada)
            refresh_tabela_total()

    elif event == '-CANCELA-SAIDA-':
        window['sai_descr'].update(value='')
        window['sai_qt'].update(value='')
        window['sai_id'].update(value='')
        window['sai_resp'].update(value='')
        window['sai_numser'].update(value='')

    # ----------------------------------------------------------------
    # ABA “Cadastro”
    # ----------------------------------------------------------------
    elif event == '-BOTAO-CADASTRO-':
        desc = values.get('cad_descr')
        try:
            qt = int(values.get('cad_qt'))
        except (ValueError, TypeError):
            sg.popup_error("Informe uma quantidade inicial válida (número inteiro).")
            continue
        num_pedido = values.get('cad_pedido')
        area_cc    = values.get('cad_area')
        if not desc:
            sg.popup_error("Informe a descrição do equipamento.")
        else:
            cadastrar_equipamento(usuario_logado, desc, qt, num_pedido, area_cc)
            # Atualiza combos e tabela
            descricoes_combo_entrada = carregar_lista_descricoes()
            descricoes_combo_saida   = carregar_lista_descricoes()
            window['ent_descr'].update(values=descricoes_combo_entrada)
            window['sai_descr'].update(values=descricoes_combo_saida)
            window['del_descr'].update(values=descricoes_combo_entrada)
            refresh_tabela_total()

    elif event == '-CANCELA-CADASTRO-':
        window['cad_descr'].update(value='')
        window['cad_qt'].update(value='')
        window['cad_pedido'].update(value='')
        window['cad_area'].update(value='')

    # ----------------------------------------------------------------
    # ABA “Exclusão”
    # ----------------------------------------------------------------
    elif event == '-BOTAO-DEL-':
        selecionado = values.get('del_descr')
        if not selecionado:
            sg.popup_error("Selecione um equipamento para deletar.")
        else:
            deletar_equipamento(usuario_logado, selecionado[0])
            # Atualiza combos e tabela
            descricoes_combo_entrada = carregar_lista_descricoes()
            descricoes_combo_saida   = carregar_lista_descricoes()
            window['ent_descr'].update(values=descricoes_combo_entrada)
            window['sai_descr'].update(values=descricoes_combo_saida)
            window['del_descr'].update(values=descricoes_combo_entrada)
            refresh_tabela_total()

    elif event == '-CANCELA-DEL-':
        window['del_descr'].update(value=[])

    # ----------------------------------------------------------------
    # ABA “Registros”
    # ----------------------------------------------------------------
    elif event == '-BOTAO-REG-ENT-':
        consultar_registros('entradas')
    elif event == '-BOTAO-REG-SAI-':
        consultar_registros('saidas')

# ----------------------------------------------------------------
# FINALIZAÇÃO
# ----------------------------------------------------------------
window.close()
